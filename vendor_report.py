"""
廠商情報報告腳本
讀取 vendors.json 和 廠商清單.txt，用 Gemini API + Google Search 搜尋各廠商近期動態，
輸出 Excel（三工作表）和 PDF 報告至 reports/{日期}/ 目錄。
"""
import json
import os
import sys
import time
from datetime import date, datetime

import anthropic
from google import genai
from google.genai import types
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def _register_cjk_font():
    """嘗試載入系統 CJK 字型，失敗則回退到 Helvetica。"""
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",  # Ubuntu GCP
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
        "C:/Windows/Fonts/msjh.ttc",   # Windows 微軟正黑體
        "C:/Windows/Fonts/msyh.ttc",   # Windows 微軟雅黑
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("CJK", path))
                # Bold：優先找同目錄的 Bold 版，找不到就沿用 Regular
                bold_path = path.replace("Regular", "Bold")
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont("CJK-Bold", bold_path))
                else:
                    pdfmetrics.registerFont(TTFont("CJK-Bold", path))
                return "CJK", "CJK-Bold"
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"

PDF_FONT, PDF_FONT_BOLD = _register_cjk_font()

load_dotenv()

# ── 路徑設定 ──
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
VENDOR_DIR   = SCRIPT_DIR
VENDOR_LIST  = os.path.join(SCRIPT_DIR, "廠商清單.txt")
VENDORS_JSON = os.path.join(SCRIPT_DIR, "vendors.json")
REPORTS_DIR  = os.path.join(SCRIPT_DIR, "reports")

TODAY = date.today().isoformat()
OUTPUT_DIR = os.path.join(REPORTS_DIR, TODAY)

# ── Gemini 客戶端 ──
gemini_client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))

# ── Anthropic Claude 客戶端（備援）──
_anthropic_client = None

def _get_anthropic_client() -> anthropic.Anthropic:
    """Lazy-init Anthropic client；未設定 API Key 時拋出明確錯誤。"""
    global _anthropic_client
    if _anthropic_client is None:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            raise RuntimeError("ANTHROPIC_API_KEY 未設定，無法啟用 Claude 備援")
        _anthropic_client = anthropic.Anthropic(api_key=api_key)
    return _anthropic_client

CLAUDE_FALLBACK_MODEL       = "claude-3-5-haiku-20241022"
CLAUDE_INPUT_COST_PER_MTOK  = 0.80   # USD per million input tokens
CLAUDE_OUTPUT_COST_PER_MTOK = 4.00   # USD per million output tokens

# 執行期狀態：一旦為 True，本次執行全程跳過 Gemini，直接用 Claude
_gemini_quota_exhausted = False

# 費用追蹤（每次執行重置，不跨執行累積）
_claude_cost_state: dict = {
    "total_input_tokens":  0,
    "total_output_tokens": 0,
    "total_cost_usd":      0.0,
    "call_count":          0,
    "limit_usd":           0.20,   # 由 load_vendor_config() 從 vendors.json 覆寫
    "limit_reached":       False,
}

SEARCH_SYSTEM = """你是成人用品產業的情報分析師，使用繁體中文（zh-TW）。
你的任務是針對指定廠商搜尋過去30天內「首次發布」的新品、停產、價格異動或重要公告。

搜尋規則（嚴格遵守）：
1. 【第1次搜尋】優先使用 site:{官網網址} 加上當年年份，例如 site:lelo.com new 2026
   若無官網，改搜尋：{廠商名稱} 新品 {當年年份}
2. 【第2次搜尋】用搜尋結果中找到的具體產品名稱，加上 {當年年份} 確認發布日期
3. 【第3次搜尋】台灣通路（蝦皮/PChome/紅犀牛），確認台灣售價與上市狀況

判斷「新品」的標準（嚴格）：
- 必須找到明確的發布日期或新聞，且日期在過去30天內
- 只有在官網/官方社群/正式媒體上有明確 {當年年份} 發布記錄，才標記為 ✅ 確認新品
- 若找到產品但無明確日期，或搜尋結果的頁面日期早於30天，標記為 ⚠️ 疑似非新品
- 無法確認則回傳空陣列，不要猜測

回傳格式（JSON）：
{
  "vendor": "廠商名稱",
  "new_products": [
    {
      "name": "產品名稱",
      "release_date": "YYYY-MM-DD 或 近似月份",
      "verification": "✅ 確認新品 / ⚠️ 疑似非新品(最早記錄:YYYY-MM) / ❓ 待確認",
      "product_type": "按摩棒/跳蛋/...",
      "body_part": "陰蒂/G點/...",
      "taiwan_status": "已上市/預購/未知",
      "taiwan_price_ntd": 數字或null,
      "channel": "通路名稱",
      "original_price": 數字或null,
      "currency": "USD/JPY/EUR/...",
      "source_url": "網址或null"
    }
  ],
  "other_changes": [
    {
      "type": "停產/價格調整/公司公告/促銷活動/其他",
      "item": "產品或項目名稱",
      "description": "說明",
      "change_date": "YYYY-MM-DD 或 近似月份",
      "source_url": "網址或null"
    }
  ],
  "activity_level": "活躍/一般/無異動"
}

若無新品，new_products 回傳空陣列。若無其他異動，other_changes 回傳空陣列。
只回傳 JSON，不要加任何說明文字。"""


def load_vendor_config():
    with open(VENDOR_LIST, encoding="utf-8") as f:
        vendor_names = [line.strip() for line in f if line.strip()]
    with open(VENDORS_JSON, encoding="utf-8") as f:
        config = json.load(f)
    vendor_map = {v["name"]: v for v in config.get("vendors", [])}
    cfg = config.get("config", {})
    # 將 vendors.json 中的費用上限設定同步到費用追蹤狀態
    _claude_cost_state["limit_usd"] = float(cfg.get("claude_cost_limit_usd", 0.20))
    return vendor_names, vendor_map, cfg


# ── 快照差異比較 ──

SNAPSHOT_DIR = os.path.join(SCRIPT_DIR, "snapshots")

DIFF_SYSTEM = """你是成人用品產業的情報分析師，使用繁體中文（zh-TW）。
以下是廠商官網的「上次快照」與「本次快照」的差異內容（新增行以 + 開頭，刪除行以 - 開頭）。
請根據差異判斷是否有新品上架、商品下架、或重要公告。

判斷原則：
- 只回報確實出現在 diff 中的變化，不要猜測
- 新增的產品名稱（+ 行）視為潛在新品，標記為 ✅ 確認新品
- 刪除的產品名稱（- 行）視為下架/停產
- 若 diff 內容不足以判斷，new_products 和 other_changes 均回傳空陣列

回傳格式（JSON）：
{
  "vendor": "廠商名稱",
  "new_products": [
    {
      "name": "產品名稱",
      "release_date": "本次快照日期",
      "verification": "✅ 確認新品",
      "product_type": "按摩棒/跳蛋/...",
      "body_part": "陰蒂/G點/...",
      "taiwan_status": "已上市/預購/未知",
      "taiwan_price_ntd": null,
      "channel": "官網",
      "original_price": null,
      "currency": null,
      "source_url": null
    }
  ],
  "other_changes": [
    {
      "type": "停產/價格調整/公司公告/促銷活動/其他",
      "item": "產品或項目名稱",
      "description": "說明",
      "change_date": "本次快照日期",
      "source_url": null
    }
  ],
  "activity_level": "活躍/一般/無異動"
}

只回傳 JSON，不要加任何說明文字。"""


def get_latest_snapshots(vendor_id: str) -> tuple[str, str]:
    """
    回傳 (上次快照內容, 本次快照內容)。
    若找不到兩份快照，回傳 ("", "")。
    """
    vendor_snap_dir = os.path.join(SNAPSHOT_DIR, vendor_id)
    if not os.path.isdir(vendor_snap_dir):
        return "", ""

    files = sorted([
        f for f in os.listdir(vendor_snap_dir)
        if f.endswith(".txt")
    ])

    if len(files) == 0:
        return "", ""
    elif len(files) == 1:
        # 只有一份（首次執行）：本次有，上次空
        with open(os.path.join(vendor_snap_dir, files[-1]), encoding="utf-8") as f:
            current = f.read()
        return "", current
    else:
        with open(os.path.join(vendor_snap_dir, files[-2]), encoding="utf-8") as f:
            previous = f.read()
        with open(os.path.join(vendor_snap_dir, files[-1]), encoding="utf-8") as f:
            current = f.read()
        return previous, current


def compute_diff(previous: str, current: str) -> str:
    """
    計算兩個文字快照的逐行差異，回傳 unified diff 格式字串。
    """
    import difflib
    prev_lines = previous.splitlines()
    curr_lines = current.splitlines()
    diff = list(difflib.unified_diff(
        prev_lines, curr_lines,
        lineterm="", n=2
    ))
    # 過濾掉 --- / +++ 標頭行，只保留 +/- 行
    result = [l for l in diff if l.startswith(("+", "-")) and not l.startswith(("---", "+++"))]
    return "\n".join(result)


# ── Claude 備援輔助函數 ──

def _call_claude_and_track(messages: list, system: str, max_tokens: int = 1500) -> "str | None":
    """
    統一的 Claude API 呼叫入口。
    - 呼叫後立即更新費用追蹤，並印出本次 / 累計費用。
    - 若費用上限已達，直接回傳 None（呼叫端需處理）。
    - 費用超限時設 limit_reached=True，後續呼叫同樣短路。
    """
    if _claude_cost_state["limit_reached"]:
        return None
    client = _get_anthropic_client()
    response = client.messages.create(
        model=CLAUDE_FALLBACK_MODEL,
        max_tokens=max_tokens,
        system=system,
        messages=messages,
    )
    usage = response.usage
    cost = (
        usage.input_tokens  * CLAUDE_INPUT_COST_PER_MTOK +
        usage.output_tokens * CLAUDE_OUTPUT_COST_PER_MTOK
    ) / 1_000_000
    _claude_cost_state["total_input_tokens"]  += usage.input_tokens
    _claude_cost_state["total_output_tokens"] += usage.output_tokens
    _claude_cost_state["total_cost_usd"]      += cost
    _claude_cost_state["call_count"]          += 1
    cumulative = _claude_cost_state["total_cost_usd"]
    limit      = _claude_cost_state["limit_usd"]
    print(
        f"    [Claude費用] 本次 ${cost:.4f} USD"
        f"（累計 ${cumulative:.4f} / 上限 ${limit:.2f} USD）",
        flush=True,
    )
    if cumulative >= limit:
        print(
            f"\n  ⚠️  Claude API 累計費用 ${cumulative:.4f} USD 已達上限 ${limit:.2f} USD，"
            f"後續廠商停止呼叫 Claude。",
            flush=True,
        )
        _claude_cost_state["limit_reached"] = True
    return response.content[0].text


def _claude_analyze_diff(vendor_name: str, diff_text: str) -> dict:
    """Claude 備援：分析網頁快照 diff（純文字分析，無工具）。"""
    if _claude_cost_state["limit_reached"]:
        return {
            "vendor": vendor_name, "new_products": [], "other_changes": [],
            "activity_level": "無異動", "_error": "Claude費用上限已達，跳過",
        }
    print(f"    [Claude備援] analyze_diff: {vendor_name}", flush=True)
    user_msg = (
        f"廠商：{vendor_name}\n"
        f"快照日期：{TODAY}\n\n"
        f"以下是本次與上次官網快照的差異（+ 為新增，- 為刪除）：\n\n"
        f"{diff_text[:3000]}\n\n"
        f"請分析差異，回傳指定 JSON 格式。"
    )
    try:
        text = _call_claude_and_track(
            messages=[{"role": "user", "content": user_msg}],
            system=DIFF_SYSTEM,
            max_tokens=1200,
        )
        if text is None:
            return {
                "vendor": vendor_name, "new_products": [], "other_changes": [],
                "activity_level": "無異動", "_error": "Claude費用上限已達，跳過",
            }
        if "```" in text:
            inner = text.split("```")[1]
            text = inner.lstrip("json").strip()
        data = json.loads(text)
        oc = data.get("other_changes", [])
        if oc and isinstance(oc[0], str):
            data["other_changes"] = [
                {"type": "其他", "item": "", "description": s,
                 "change_date": TODAY, "source_url": None}
                for s in oc
            ]
        return data
    except Exception as e:
        print(f"  [Claude Diff錯誤] {vendor_name}: {e}", file=sys.stderr)
        return {
            "vendor": vendor_name, "new_products": [], "other_changes": [],
            "activity_level": "無異動", "_error": str(e),
        }


def _claude_search_vendor(vendor_name: str, vendor_info: dict) -> dict:
    """
    Claude 備援：預先抓取廠商官網頁面，再請 Claude 分析。
    因 Claude 無內建 Google Search，改以 snapshot_fetcher.fetch_page_text 抓取靜態頁面。
    """
    if _claude_cost_state["limit_reached"]:
        return {
            "vendor": vendor_name, "new_products": [], "other_changes": [],
            "activity_level": "無異動", "_error": "Claude費用上限已達，跳過",
        }

    import snapshot_fetcher as _sf  # noqa: PLC0415（模組已在 main 中載入，此為複用）

    official_site  = vendor_info.get("official_site") or ""
    snapshot_pages = vendor_info.get("snapshot_pages", [])

    # 建立抓取候選 URL（最多 3 個，避免耗時過長）
    fetch_urls: list = list(snapshot_pages)
    if official_site:
        base = official_site.rstrip("/")
        for suffix in ["/new-arrivals", "/news", "/what-s-new", "/products/new", "/"]:
            candidate = base + suffix
            if candidate not in fetch_urls:
                fetch_urls.append(candidate)
                break

    fetched_sections: list = []
    print(f"    [Claude備援] search_vendor: {vendor_name}，抓取官網頁面 ...", flush=True)
    for url in fetch_urls[:3]:
        print(f"      抓取: {url}", flush=True)
        page_text = _sf.fetch_page_text(url)
        if page_text:
            fetched_sections.append(f"=== {url} ===\n{page_text[:2000]}")
        time.sleep(1)

    year   = date.today().year
    region = vendor_info.get("region", "未知")
    notes  = vendor_info.get("notes", "")

    if fetched_sections:
        user_msg = (
            f"廠商：{vendor_name}\n地區：{region}\n"
            f"官網：{official_site or '無'}\n備註：{notes}\n\n"
            f"以下是從該廠商官網抓取的頁面內容（截至 {TODAY}）：\n\n"
            + "\n\n".join(fetched_sections)[:5000]
            + f"\n\n只回報有明確 {year} 年日期證據的新品。"
              f"無法確認則 new_products 回傳空陣列。回傳指定 JSON 格式。"
        )
    else:
        user_msg = (
            f"廠商：{vendor_name}\n地區：{region}\n官網：{official_site or '無'}\n備註：{notes}\n\n"
            f"注意：無法取得該廠商官網頁面內容（可能需要 JS 渲染或頁面不可存取）。\n"
            f"請回傳 JSON 格式，new_products 和 other_changes 均設為空陣列，"
            f"activity_level 設為「無異動」。"
        )

    try:
        text = _call_claude_and_track(
            messages=[{"role": "user", "content": user_msg}],
            system=SEARCH_SYSTEM,
            max_tokens=1500,
        )
        if text is None:
            return {
                "vendor": vendor_name, "new_products": [], "other_changes": [],
                "activity_level": "無異動", "_error": "Claude費用上限已達，跳過",
            }
        if "```" in text:
            inner = text.split("```")[1]
            text = inner.lstrip("json").strip()
        data = json.loads(text)
        # 正規化 activity_level（Claude 偶爾回傳英文）
        al = data.get("activity_level", "")
        al_map = {
            "active": "活躍", "high": "活躍",
            "medium": "一般", "moderate": "一般", "normal": "一般",
            "low": "無異動", "none": "無異動", "inactive": "無異動",
        }
        if al not in ("活躍", "一般", "無異動"):
            data["activity_level"] = al_map.get(al.lower(), "一般")
        oc = data.get("other_changes", [])
        if oc and isinstance(oc[0], str):
            data["other_changes"] = [
                {"type": "其他", "item": "", "description": s,
                 "change_date": TODAY, "source_url": None}
                for s in oc
            ]
        return data
    except Exception as e:
        print(f"  [Claude Search錯誤] {vendor_name}: {e}", file=sys.stderr)
        return {
            "vendor": vendor_name, "new_products": [], "other_changes": [],
            "activity_level": "無異動", "_error": str(e),
        }


def analyze_diff(vendor_name: str, diff_text: str) -> dict:
    """將 diff 文字送給 Gemini Flash 分析，回傳結構化 JSON。配額耗盡時備援至 Claude。"""
    global _gemini_quota_exhausted

    # 若本次執行已切換至 Claude，直接跳過 Gemini
    if _gemini_quota_exhausted:
        return _claude_analyze_diff(vendor_name, diff_text)

    user_msg = (
        f"廠商：{vendor_name}\n"
        f"快照日期：{TODAY}\n\n"
        f"以下是本次與上次官網快照的差異（+ 為新增，- 為刪除）：\n\n"
        f"{diff_text[:3000]}\n\n"
        f"請分析差異，回傳指定 JSON 格式。"
    )
    for attempt in range(3):
        try:
            response = gemini_client.models.generate_content(
                model="gemini-2.5-flash",
                contents=user_msg,
                config=types.GenerateContentConfig(
                    system_instruction=DIFF_SYSTEM,
                    response_mime_type="application/json",
                    max_output_tokens=1000,
                    response_schema={
                        "type": "object",
                        "properties": {
                            "vendor": {"type": "string"},
                            "new_products": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "name": {"type": "string"},
                                        "release_date": {"type": "string"},
                                        "verification": {"type": "string"},
                                        "product_type": {"type": "string"},
                                        "body_part": {"type": "string"},
                                        "taiwan_status": {"type": "string"},
                                        "taiwan_price_ntd": {"type": "number", "nullable": True},
                                        "channel": {"type": "string"},
                                        "original_price": {"type": "number", "nullable": True},
                                        "currency": {"type": "string", "nullable": True},
                                        "source_url": {"type": "string", "nullable": True},
                                    }
                                }
                            },
                            "other_changes": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "type": {"type": "string"},
                                        "item": {"type": "string"},
                                        "description": {"type": "string"},
                                        "change_date": {"type": "string"},
                                        "source_url": {"type": "string", "nullable": True},
                                    }
                                }
                            },
                            "activity_level": {
                                "type": "string",
                                "enum": ["活躍", "一般", "無異動"]
                            },
                        },
                        "required": ["vendor", "new_products", "other_changes", "activity_level"]
                    },
                ),
            )
            data = json.loads(response.text)
            # 正規化 other_changes（若為字串陣列則轉物件）
            oc = data.get("other_changes", [])
            if oc and isinstance(oc[0], str):
                data["other_changes"] = [
                    {"type": "其他", "item": "", "description": s, "change_date": TODAY, "source_url": None}
                    for s in oc
                ]
            return data
        except Exception as e:
            err_str = str(e)
            quota_hit = "RESOURCE_EXHAUSTED" in err_str or "429" in err_str
            if quota_hit and attempt < 2:
                wait = 60 * (attempt + 1)
                print(f"  [Rate limit] 等待 {wait} 秒...", flush=True)
                time.sleep(wait)
            elif quota_hit:
                # 3 次均失敗且為配額問題 → 本次執行全程切換至 Claude
                print(f"  [Gemini配額耗盡] 切換至 Claude 備援（本次執行全程）", flush=True)
                _gemini_quota_exhausted = True
                break  # 跳出 retry loop，fall through 至 Claude
            else:
                print(f"  [Diff分析錯誤] {vendor_name}: {e}", file=sys.stderr)
                return {
                    "vendor": vendor_name,
                    "new_products": [],
                    "other_changes": [],
                    "activity_level": "無異動",
                    "_error": err_str,
                }

    # Gemini 配額耗盡後的備援路徑
    return _claude_analyze_diff(vendor_name, diff_text)


def search_vendor(vendor_name: str, vendor_info: dict) -> dict:
    """呼叫 Gemini API with Google Search 搜尋單一廠商資訊。配額耗盡時備援至 Claude。"""
    global _gemini_quota_exhausted

    # 若本次執行已切換至 Claude，直接跳過 Gemini
    if _gemini_quota_exhausted:
        return _claude_search_vendor(vendor_name, vendor_info)

    keywords = vendor_info.get("taiwan_search", []) + vendor_info.get("search_keywords", [])
    official_site = vendor_info.get("official_site") or ""
    region = vendor_info.get("region", "未知")
    notes = vendor_info.get("notes", "")

    year = date.today().year
    if official_site:
        # 從 URL 取出 domain（去掉 https://www. 前綴）
        domain = official_site.replace("https://", "").replace("http://", "").rstrip("/")
        first_search = f"site:{domain} new {year} OR site:{domain} 新品 {year}"
    else:
        first_search = f"{vendor_name} 新品 {year} OR {vendor_name} new product {year}"

    taiwan_keywords = ", ".join(vendor_info.get("taiwan_search", [])[:2])

    user_msg = (
        f"請搜尋廠商「{vendor_name}」在 {year} 年發布的新品與近期動態（截至 {TODAY}）。\n"
        f"地區：{region}\n"
        f"官網：{official_site or '無'}\n"
        f"備註：{notes}\n\n"
        f"請依照以下順序執行搜尋：\n"
        f"1. 官網新品（建議query：{first_search}）\n"
        f"2. 用找到的產品名稱 + {year} 確認發布日期\n"
        f"3. 台灣通路售價（建議query：{taiwan_keywords}）\n\n"
        f"只回報有明確 {year} 年日期證據的新品。無法確認則 new_products 回傳空陣列。\n"
        f"回傳指定 JSON 格式。"
    )

    for attempt in range(3):
        try:
            response = gemini_client.models.generate_content(
                model="gemini-2.5-flash",
                contents=user_msg,
                config=types.GenerateContentConfig(
                    system_instruction=SEARCH_SYSTEM,
                    tools=[types.Tool(google_search=types.GoogleSearch())],
                ),
            )
            result_text = response.text.strip()
            # 解析 JSON：移除 markdown code block 包裝
            if "```" in result_text:
                inner = result_text.split("```")[1]
                result_text = inner.lstrip("json").strip()
            data = json.loads(result_text)
            # 正規化 activity_level（Gemini 可能回傳英文）
            al = data.get("activity_level", "")
            al_map = {
                "active": "活躍", "high": "活躍",
                "medium": "一般", "moderate": "一般", "normal": "一般",
                "low": "無異動", "none": "無異動", "inactive": "無異動",
            }
            if al not in ("活躍", "一般", "無異動"):
                data["activity_level"] = al_map.get(al.lower(), "一般")
            # 正規化 other_changes（若為字串陣列則轉物件）
            oc = data.get("other_changes", [])
            if oc and isinstance(oc[0], str):
                data["other_changes"] = [
                    {"type": "其他", "item": "", "description": s, "change_date": TODAY, "source_url": None}
                    for s in oc
                ]
            return data
        except Exception as e:
            err_str = str(e)
            quota_hit = "RESOURCE_EXHAUSTED" in err_str or "429" in err_str
            if quota_hit and attempt < 2:
                wait = 60 * (attempt + 1)
                print(f"  [Rate limit] {vendor_name}，等待 {wait} 秒後重試（第 {attempt+1} 次）...", flush=True)
                time.sleep(wait)
            elif quota_hit:
                # 3 次均失敗且為配額問題 → 本次執行全程切換至 Claude
                print(f"  [Gemini配額耗盡] 切換至 Claude 備援（本次執行全程）", flush=True)
                _gemini_quota_exhausted = True
                break  # 跳出 retry loop，fall through 至 Claude
            else:
                print(f"  [錯誤] {vendor_name}: {e}", file=sys.stderr)
                return {
                    "vendor": vendor_name,
                    "new_products": [],
                    "other_changes": [],
                    "activity_level": "無異動",
                    "_error": err_str,
                }

    # Gemini 配額耗盡後的備援路徑
    return _claude_search_vendor(vendor_name, vendor_info)


# ── 活躍度分析常數 ──
LEVEL_ORDER = {"": 0, "無異動": 0, "一般": 1, "活躍": 2}
LEVEL_LABEL = {0: "無異動", 1: "一般", 2: "活躍"}
EXCHANGE_RATES = {"USD": 32, "JPY": 0.21, "EUR": 35, "GBP": 40, "CNY": 4.4, "AUD": 20}

# ── Excel 生成 ──

HEADER_FILL = PatternFill("solid", fgColor="ADD8E6")  # 淺藍
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")  # 淺灰
BOLD        = Font(name="微軟正黑體", bold=True)
NORMAL      = Font(name="微軟正黑體")

TAB_COLORS = {"overview": "4472C4", "new_products": "70AD47", "other_changes": "ED7D31"}


# ── 歷史資料讀取 ──

def load_history_data() -> dict:
    """
    掃描 REPORTS_DIR 底下所有歷史 xlsx，讀取「每週總覽」與「新品詳細」兩個 sheet。
    排除今天（TODAY）的資料夾（避免與本週新結果重複）。
    回傳 {"overview": [...], "new_products": [...]}
    """
    import re
    overview_rows = []
    product_rows  = []

    if not os.path.isdir(REPORTS_DIR):
        return {"overview": overview_rows, "new_products": product_rows}

    folders = sorted([
        f for f in os.listdir(REPORTS_DIR)
        if os.path.isdir(os.path.join(REPORTS_DIR, f))
        and re.match(r'^\d{4}-\d{2}-\d{2}$', f)
        and f != TODAY
    ])

    for folder in folders:
        folder_path = os.path.join(REPORTS_DIR, folder)
        xlsx_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
        if not xlsx_files:
            continue
        xlsx_path = os.path.join(folder_path, xlsx_files[0])
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, data_only=True)

            # 每週總覽 sheet
            if "每週總覽" in wb.sheetnames:
                ws = wb["每週總覽"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= 2:
                    headers = [str(h) if h is not None else "" for h in rows[0]]
                    for row in rows[1:]:
                        if row[0] is None:
                            continue
                        d = dict(zip(headers, row))
                        d["週次"] = folder
                        overview_rows.append(d)

            # 新品詳細 sheet
            if "新品詳細" in wb.sheetnames:
                ws = wb["新品詳細"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= 2:
                    headers = [str(h) if h is not None else "" for h in rows[0]]
                    for row in rows[1:]:
                        if row[0] is None or str(row[0]) == "本週無新品發布":
                            continue
                        d = dict(zip(headers, row))
                        d["週次"] = folder
                        product_rows.append(d)
        except Exception as e:
            print(f"  [歷史資料] 讀取 {xlsx_path} 失敗，跳過：{e}")

    print(f"  歷史資料讀取完成：{len(folders)} 個資料夾，{len(overview_rows)} 筆總覽，{len(product_rows)} 筆新品")
    return {"overview": overview_rows, "new_products": product_rows}


def _get_field(row, *keys):
    """嘗試多個 key 名稱（中英文），回傳第一個非 None 值。"""
    for k in keys:
        v = row.get(k)
        if v is not None and v != "":
            return v
    return None


# ── HTML 分析資料準備 ──

def build_trend_data(results: list, history: dict) -> tuple:
    """
    整合歷史 overview + 本週 results，建立活躍度趨勢矩陣。
    回傳 (trend_data_for_html, alert_list)
    """
    # 建立 matrix[廠商][週次] = level
    from collections import defaultdict
    matrix = defaultdict(dict)

    # 歷史資料
    for row in history.get("overview", []):
        vendor = _get_field(row, "廠商名稱", "vendor") or ""
        week   = row.get("週次", "")
        level_str = _get_field(row, "整體動態評估", "activity_level") or ""
        if vendor and week:
            matrix[vendor][week] = LEVEL_ORDER.get(level_str, 0)

    # 本週資料
    for r in results:
        vendor = r.get("vendor", "")
        level_str = r.get("activity_level", "無異動")
        if vendor:
            matrix[vendor][TODAY] = LEVEL_ORDER.get(level_str, 0)

    all_vendors = sorted(matrix.keys())
    all_weeks   = sorted(set(w for v in matrix.values() for w in v))

    # 建立 JS 用的 datasets
    CHART_COLORS = [
        "#e6194b","#3cb44b","#4363d8","#f58231","#911eb4",
        "#42d4f4","#f032e6","#bfef45","#fabed4","#469990",
        "#dcbeff"
    ]
    datasets = []
    for i, vendor in enumerate(all_vendors):
        data = [matrix[vendor].get(w, None) for w in all_weeks]
        datasets.append({
            "label": vendor,
            "data": data,
            "borderColor": CHART_COLORS[i % len(CHART_COLORS)],
            "backgroundColor": CHART_COLORS[i % len(CHART_COLORS)],
            "tension": 0.3,
            "fill": False,
            "spanGaps": True,
        })

    trend_data = {
        "labels": all_weeks,
        "datasets": datasets,
    }

    # 偵測活躍度上升（比較最後兩週）
    alerts = []
    if len(all_weeks) >= 2:
        prev_week = all_weeks[-2]
        curr_week = all_weeks[-1]
        if curr_week == TODAY:
            for vendor in all_vendors:
                prev_level = matrix[vendor].get(prev_week, 0)
                curr_level = matrix[vendor].get(curr_week, 0)
                if curr_level > prev_level:
                    alerts.append(vendor)

    return trend_data, alerts


def build_matrix_data(results: list, history: dict) -> dict:
    """
    建立產品類型 × 施用部位 的競爭矩陣。
    只計算 verification == "✅ 確認新品" 的產品。
    """
    from collections import defaultdict

    # 收集所有確認新品
    all_products = []

    # 歷史資料
    for row in history.get("new_products", []):
        verif = _get_field(row, "新品核實", "verification") or ""
        if "確認新品" in verif:
            all_products.append({
                "body_part":    _get_field(row, "施用部位", "body_part") or "未分類",
                "product_type": _get_field(row, "產品類型", "product_type") or "未分類",
                "vendor":       _get_field(row, "廠商", "vendor") or "",
                "name":         _get_field(row, "產品名稱", "name") or "",
            })

    # 本週資料
    for r in results:
        for p in r.get("new_products", []):
            if "確認新品" in (p.get("verification") or ""):
                all_products.append({
                    "body_part":    p.get("body_part") or "未分類",
                    "product_type": p.get("product_type") or "未分類",
                    "vendor":       r.get("vendor", ""),
                    "name":         p.get("name", ""),
                })

    if not all_products:
        return {"rows": [], "cols": [], "matrix": [], "products": {}}

    # 建立 cross-tab
    count_map   = defaultdict(lambda: defaultdict(int))
    product_map = defaultdict(list)
    all_bodies  = set()
    all_types   = set()

    for p in all_products:
        bodies = [x.strip() for x in str(p["body_part"]).split("/") if x.strip()]
        types  = [x.strip() for x in str(p["product_type"]).split("/") if x.strip()]
        for b in bodies:
            for t in types:
                count_map[b][t] += 1
                product_map[f"{b}||{t}"].append({"廠商": p["vendor"], "名稱": p["name"]})
                all_bodies.add(b)
                all_types.add(t)

    row_labels = sorted(all_bodies)
    col_labels = sorted(all_types)
    matrix = [[count_map[b][t] for t in col_labels] for b in row_labels]

    return {
        "rows": row_labels,
        "cols": col_labels,
        "matrix": matrix,
        "products": {k: v for k, v in product_map.items()},
    }


def build_pricing_data(results: list, history: dict) -> dict:
    """
    建立定價分析資料：各廠商台灣售價範圍 + 換算倍率散佈圖。
    """
    from collections import defaultdict

    all_items = []

    def _parse_price(val):
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    # 歷史資料
    for row in history.get("new_products", []):
        tw  = _parse_price(_get_field(row, "台灣售價(NTD)", "taiwan_price_ntd"))
        orig = _parse_price(_get_field(row, "原廠售價", "original_price"))
        curr = str(_get_field(row, "原廠幣別", "currency") or "").upper()
        vendor = _get_field(row, "廠商", "vendor") or "未知"
        name   = _get_field(row, "產品名稱", "name") or ""
        if tw or orig:
            ratio = None
            if tw and orig and curr in EXCHANGE_RATES:
                orig_ntd = orig * EXCHANGE_RATES[curr]
                if orig_ntd > 0:
                    ratio = round(tw / orig_ntd, 2)
            all_items.append({"vendor": vendor, "name": name, "tw": tw, "orig": orig, "curr": curr, "ratio": ratio})

    # 本週資料
    for r in results:
        for p in r.get("new_products", []):
            tw   = _parse_price(p.get("taiwan_price_ntd"))
            orig = _parse_price(p.get("original_price"))
            curr = str(p.get("currency") or "").upper()
            vendor = r.get("vendor", "未知")
            name   = p.get("name", "")
            if tw or orig:
                ratio = None
                if tw and orig and curr in EXCHANGE_RATES:
                    orig_ntd = orig * EXCHANGE_RATES[curr]
                    if orig_ntd > 0:
                        ratio = round(tw / orig_ntd, 2)
                all_items.append({"vendor": vendor, "name": name, "tw": tw, "orig": orig, "curr": curr, "ratio": ratio})

    if not all_items:
        return {"vendors": [], "ntd_min": [], "ntd_max": [], "ntd_avg": [], "scatter_pts": []}

    # 按廠商彙整
    vendor_data = defaultdict(list)
    for item in all_items:
        if item["tw"]:
            vendor_data[item["vendor"]].append(item["tw"])

    vendors  = sorted(vendor_data.keys())
    ntd_min  = [min(vendor_data[v]) for v in vendors]
    ntd_max  = [max(vendor_data[v]) for v in vendors]
    ntd_avg  = [round(sum(vendor_data[v]) / len(vendor_data[v]), 0) for v in vendors]

    # 散佈圖：換算倍率
    scatter_pts = []
    for item in all_items:
        if item["ratio"] is not None and item["vendor"] in vendors:
            x_idx = vendors.index(item["vendor"])
            scatter_pts.append({
                "x": x_idx,
                "y": item["ratio"],
                "label": f"{item['vendor']} - {item['name']}",
            })

    return {
        "vendors": vendors,
        "ntd_min":  ntd_min,
        "ntd_max":  ntd_max,
        "ntd_avg":  ntd_avg,
        "scatter_pts": scatter_pts,
    }


# ── HTML 互動分析報告 ──

def build_html(output_path: str, trend_data: dict, matrix_data: dict, pricing_data: dict, results: list):
    """產生單一自包含 HTML 互動分析報告（Chart.js via CDN）。"""
    import json as _json

    vendor_count = len({r.get("vendor","") for r in results if r.get("vendor")})

    trend_json   = _json.dumps(trend_data,   ensure_ascii=False)
    matrix_json  = _json.dumps(matrix_data,  ensure_ascii=False)
    pricing_json = _json.dumps(pricing_data, ensure_ascii=False)

    # 產品矩陣熱力圖的色階（值 0→灰, 1→淺藍, 2→中藍, ≥3→深藍）
    def _cell_color(val):
        if val == 0:   return "#eeeeee"
        elif val == 1: return "#b3d9f5"
        elif val == 2: return "#5baee0"
        else:          return "#1565c0"

    # 建立 matrix HTML table
    if matrix_data["rows"]:
        thead = "<tr><th>施用部位 \\ 產品類型</th>" + "".join(f"<th>{c}</th>" for c in matrix_data["cols"]) + "<th>合計</th></tr>"
        tbody_rows = []
        for ri, row_label in enumerate(matrix_data["rows"]):
            row_vals = matrix_data["matrix"][ri]
            row_total = sum(row_vals)
            cells = ""
            for ci, val in enumerate(row_vals):
                key = f"{row_label}||{matrix_data['cols'][ci]}"
                color = _cell_color(val)
                if val > 0:
                    cells += f'<td style="background:{color};cursor:pointer;" onclick="showModal(\'{key}\')">{val}</td>'
                else:
                    cells += f'<td style="background:{color};">0</td>'
            tbody_rows.append(f"<tr><td class='row-label'>{row_label}</td>{cells}<td class='total'>{row_total}</td></tr>")
        col_totals = [sum(matrix_data["matrix"][ri][ci] for ri in range(len(matrix_data["rows"]))) for ci in range(len(matrix_data["cols"]))]
        total_row = "<tr class='total-row'><td>合計</td>" + "".join(f"<td>{v}</td>" for v in col_totals) + f"<td>{sum(col_totals)}</td></tr>"
        matrix_table = f"<table class='matrix-table'><thead>{thead}</thead><tbody>{''.join(tbody_rows)}{total_row}</tbody></table>"
    else:
        matrix_table = "<p class='no-data'>尚無確認新品資料，下次執行後開始累積。</p>"

    has_pricing  = bool(pricing_data["vendors"])
    has_scatter  = bool(pricing_data["scatter_pts"])

    html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>廠商情報分析 — {TODAY}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #f4f6f9; --card: #ffffff; --text: #1a1a2e; --accent: #1565c0;
    --border: #dde3ed; --muted: #6b7280;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft JhengHei", "Noto Sans TC", sans-serif; background: var(--bg); color: var(--text); padding: 24px; }}
  h1 {{ font-size: 1.5rem; font-weight: 700; color: var(--accent); margin-bottom: 4px; }}
  .subtitle {{ color: var(--muted); font-size: 0.875rem; margin-bottom: 24px; }}
  .card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 24px; margin-bottom: 24px; box-shadow: 0 1px 4px rgba(0,0,0,.06); }}
  .card h2 {{ font-size: 1.1rem; color: var(--accent); margin-bottom: 16px; border-bottom: 2px solid var(--border); padding-bottom: 8px; }}
  .chart-wrap {{ position: relative; height: 340px; }}
  .chart-wrap-sm {{ position: relative; height: 280px; }}
  .matrix-table {{ border-collapse: collapse; width: 100%; font-size: 0.8rem; }}
  .matrix-table th, .matrix-table td {{ border: 1px solid var(--border); padding: 6px 10px; text-align: center; }}
  .matrix-table th {{ background: #1565c0; color: #fff; white-space: nowrap; }}
  .matrix-table .row-label {{ text-align: left; font-weight: 600; background: #e8f0fe; }}
  .matrix-table .total {{ font-weight: 700; background: #e8f0fe; }}
  .matrix-table .total-row td {{ font-weight: 700; background: #d0e0fc; }}
  .no-data {{ color: var(--muted); font-style: italic; padding: 12px 0; }}
  .legend-bar {{ display: flex; gap: 16px; flex-wrap: wrap; margin-top: 12px; font-size: 0.78rem; }}
  .legend-item {{ display: flex; align-items: center; gap: 6px; }}
  .legend-swatch {{ width: 18px; height: 18px; border-radius: 3px; border: 1px solid #ccc; }}
  /* Modal */
  .modal-overlay {{ display:none; position:fixed; inset:0; background:rgba(0,0,0,.45); z-index:1000; justify-content:center; align-items:center; }}
  .modal-overlay.show {{ display:flex; }}
  .modal {{ background:#fff; border-radius:12px; padding:24px; max-width:480px; width:90%; max-height:80vh; overflow-y:auto; }}
  .modal h3 {{ margin-bottom:12px; color:var(--accent); }}
  .modal ul {{ padding-left:18px; }}
  .modal ul li {{ margin-bottom:6px; font-size:0.9rem; }}
  .modal-close {{ margin-top:16px; padding:8px 20px; background:var(--accent); color:#fff; border:none; border-radius:6px; cursor:pointer; font-size:0.9rem; }}
  /* Y軸標籤說明 */
  .y-legend {{ font-size: 0.75rem; color: var(--muted); margin-bottom: 8px; }}
  @media (max-width: 600px) {{
    .chart-wrap {{ height: 260px; }}
    .matrix-table {{ font-size: 0.72rem; }}
  }}
</style>
</head>
<body>
<h1>廠商情報分析報告</h1>
<p class="subtitle">報告日期：{TODAY}　｜　搜尋區間：過去 30 天　｜　監測廠商：{vendor_count}家</p>

<!-- Section 1: 活躍度趨勢 -->
<div class="card">
  <h2>一、廠商活躍度趨勢</h2>
  <p class="y-legend">Y 軸：0 = 無異動　1 = 一般　2 = 活躍</p>
  <div class="chart-wrap">
    <canvas id="trendChart"></canvas>
  </div>
</div>

<!-- Section 2: 產品矩陣 -->
<div class="card">
  <h2>二、產品類型競爭矩陣（累計確認新品數）</h2>
  <div class="legend-bar">
    <span class="legend-item"><span class="legend-swatch" style="background:#eeeeee"></span>0（無競爭）</span>
    <span class="legend-item"><span class="legend-swatch" style="background:#b3d9f5"></span>1 件</span>
    <span class="legend-item"><span class="legend-swatch" style="background:#5baee0"></span>2 件</span>
    <span class="legend-item"><span class="legend-swatch" style="background:#1565c0;border-color:#1565c0"></span>≥3 件（激烈）</span>
  </div>
  <div style="overflow-x:auto; margin-top:12px;">
    {matrix_table}
  </div>
</div>

<!-- Section 3: 定價分析 -->
<div class="card">
  <h2>三、台灣售價區間比較（NTD）</h2>
  {'<div class="chart-wrap-sm"><canvas id="pricingChart"></canvas></div>' if has_pricing else '<p class="no-data">尚無定價資料，需有台灣售價紀錄後才顯示。</p>'}
  <h2 style="margin-top:20px;">換算倍率分布（台灣售價 ÷ 原廠售價×參考匯率）</h2>
  <p class="y-legend">參考匯率：USD×32、JPY×0.21、EUR×35、GBP×40、CNY×4.4、AUD×20。倍率＞1 表示台灣售價高於原廠換算。</p>
  {'<div class="chart-wrap-sm"><canvas id="scatterChart"></canvas></div>' if has_scatter else '<p class="no-data">尚無換算倍率資料（需同時有原廠售價與台灣售價）。</p>'}
</div>

<!-- Modal -->
<div class="modal-overlay" id="modalOverlay">
  <div class="modal">
    <h3 id="modalTitle"></h3>
    <ul id="modalList"></ul>
    <button class="modal-close" onclick="closeModal()">關閉</button>
  </div>
</div>

<script>
const TREND_DATA   = {trend_json};
const MATRIX_DATA  = {matrix_json};
const PRICING_DATA = {pricing_json};

// ── 折線圖 ──
(function() {{
  const ctx = document.getElementById('trendChart');
  if (!ctx) return;
  new Chart(ctx, {{
    type: 'line',
    data: TREND_DATA,
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      scales: {{
        y: {{
          min: 0, max: 2,
          ticks: {{
            stepSize: 1,
            callback: v => ['無異動','一般','活躍'][v] || v,
          }},
          grid: {{ color: '#eee' }},
        }},
        x: {{ grid: {{ color: '#eee' }} }},
      }},
      plugins: {{
        tooltip: {{
          callbacks: {{
            label: ctx => ctx.dataset.label + '：' + (['無異動','一般','活躍'][ctx.parsed.y] || ctx.parsed.y),
          }},
        }},
        legend: {{ position: 'bottom' }},
      }},
    }},
  }});
}})();

// ── 定價長條圖 ──
(function() {{
  const ctx = document.getElementById('pricingChart');
  if (!ctx || !PRICING_DATA.vendors.length) return;
  new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels: PRICING_DATA.vendors,
      datasets: [
        {{ label: '最低售價', data: PRICING_DATA.ntd_min, backgroundColor: 'rgba(91,174,224,0.7)', borderColor: '#5baee0', borderWidth: 1 }},
        {{ label: '平均售價', data: PRICING_DATA.ntd_avg, backgroundColor: 'rgba(21,101,192,0.8)', borderColor: '#1565c0', borderWidth: 1 }},
        {{ label: '最高售價', data: PRICING_DATA.ntd_max, backgroundColor: 'rgba(179,217,245,0.7)', borderColor: '#b3d9f5', borderWidth: 1 }},
      ],
    }},
    options: {{
      indexAxis: 'y',
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{ position: 'bottom' }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label + '：NT$' + c.parsed.x.toLocaleString() }} }},
      }},
      scales: {{
        x: {{ ticks: {{ callback: v => 'NT$' + v.toLocaleString() }} }},
      }},
    }},
  }});
}})();

// ── 換算倍率散佈圖 ──
(function() {{
  const ctx = document.getElementById('scatterChart');
  if (!ctx || !PRICING_DATA.scatter_pts.length) return;
  new Chart(ctx, {{
    type: 'scatter',
    data: {{
      datasets: [{{
        label: '換算倍率',
        data: PRICING_DATA.scatter_pts,
        backgroundColor: 'rgba(21,101,192,0.6)',
        pointRadius: 6,
        pointHoverRadius: 8,
      }}],
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      scales: {{
        x: {{
          ticks: {{
            callback: v => PRICING_DATA.vendors[v] || v,
            maxRotation: 30,
          }},
          min: -0.5,
          max: PRICING_DATA.vendors.length - 0.5,
        }},
        y: {{
          title: {{ display: true, text: '換算倍率' }},
          grid: {{ color: '#eee' }},
        }},
      }},
      plugins: {{
        tooltip: {{
          callbacks: {{
            label: c => c.raw.label + '   倍率：' + c.raw.y,
          }},
        }},
        legend: {{ display: false }},
      }},
    }},
  }});
}})();

// ── Modal ──
function showModal(key) {{
  const products = MATRIX_DATA.products[key] || [];
  const parts = key.split('||');
  document.getElementById('modalTitle').textContent = parts[0] + ' × ' + parts[1] + '（' + products.length + ' 件）';
  const ul = document.getElementById('modalList');
  ul.innerHTML = '';
  products.forEach(p => {{
    const li = document.createElement('li');
    li.textContent = '[' + p['廠商'] + '] ' + p['名稱'];
    ul.appendChild(li);
  }});
  document.getElementById('modalOverlay').classList.add('show');
}}
function closeModal() {{
  document.getElementById('modalOverlay').classList.remove('show');
}}
document.getElementById('modalOverlay').addEventListener('click', e => {{
  if (e.target === document.getElementById('modalOverlay')) closeModal();
}});
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


def _set_col_widths(ws, min_w=12, max_w=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


def _write_row(ws, row_idx, values, bold=False, fill=None):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = BOLD if bold else NORMAL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def build_excel(results: list[dict], output_path: str):
    wb = Workbook()

    # ── 工作表 1：每週總覽 ──
    ws1 = wb.active
    ws1.title = "每週總覽"
    ws1.sheet_properties.tabColor = TAB_COLORS["overview"]
    headers1 = ["廠商名稱", "地區", "新品數量", "停產數量", "價格異動", "重要公告", "整體動態評估", "資料更新日期"]
    _write_row(ws1, 1, headers1, bold=True, fill=HEADER_FILL)
    for i, r in enumerate(results, 2):
        fill = ALT_FILL if i % 2 == 0 else None
        new_cnt     = len(r.get("new_products", []))
        stopped_cnt = sum(1 for c in r.get("other_changes", []) if c.get("type") == "停產")
        price_cnt   = sum(1 for c in r.get("other_changes", []) if c.get("type") == "價格調整")
        announce_cnt= sum(1 for c in r.get("other_changes", []) if c.get("type") == "公司公告")
        _write_row(ws1, i, [
            r.get("vendor", ""),
            r.get("_region", ""),
            new_cnt,
            stopped_cnt,
            price_cnt,
            announce_cnt,
            r.get("activity_level", "無異動"),
            TODAY,
        ], fill=fill)
    ws1.auto_filter.ref = ws1.dimensions
    _set_col_widths(ws1)

    # ── 工作表 2：新品詳細 ──
    ws2 = wb.create_sheet("新品詳細")
    ws2.sheet_properties.tabColor = TAB_COLORS["new_products"]
    headers2 = ["廠商", "產品名稱", "發布日期", "新品核實", "產品類型", "施用部位",
                "台灣上市狀況", "台灣售價(NTD)", "來源通路", "原廠售價", "原廠幣別", "來源網址"]
    _write_row(ws2, 1, headers2, bold=True, fill=HEADER_FILL)
    row2 = 2
    has_products = False
    for r in results:
        for p in r.get("new_products", []):
            fill = ALT_FILL if row2 % 2 == 0 else None
            _write_row(ws2, row2, [
                r.get("vendor", ""),
                p.get("name", ""),
                p.get("release_date", ""),
                p.get("verification", "❓ 待確認"),
                p.get("product_type", ""),
                p.get("body_part", ""),
                p.get("taiwan_status", ""),
                p.get("taiwan_price_ntd"),
                p.get("channel", ""),
                p.get("original_price"),
                p.get("currency", ""),
                p.get("source_url", ""),
            ], fill=fill)
            row2 += 1
            has_products = True
    if not has_products:
        ws2.cell(row=2, column=1, value="本週無新品發布").font = NORMAL
    ws2.auto_filter.ref = ws2.dimensions
    _set_col_widths(ws2)

    # ── 工作表 3：其他異動 ──
    ws3 = wb.create_sheet("其他異動")
    ws3.sheet_properties.tabColor = TAB_COLORS["other_changes"]
    headers3 = ["廠商", "異動類型", "產品/項目名稱", "異動內容說明", "異動日期", "來源網址"]
    _write_row(ws3, 1, headers3, bold=True, fill=HEADER_FILL)
    row3 = 2
    has_changes = False
    for r in results:
        for c in r.get("other_changes", []):
            fill = ALT_FILL if row3 % 2 == 0 else None
            _write_row(ws3, row3, [
                r.get("vendor", ""),
                c.get("type", ""),
                c.get("item", ""),
                c.get("description", ""),
                c.get("change_date", ""),
                c.get("source_url", ""),
            ], fill=fill)
            row3 += 1
            has_changes = True
    if not has_changes:
        ws3.cell(row=2, column=1, value="本週無其他異動").font = NORMAL
    ws3.auto_filter.ref = ws3.dimensions
    _set_col_widths(ws3)

    wb.save(output_path)


# ── PDF 生成 ──

def build_pdf(results: list[dict], output_path: str):
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontName=PDF_FONT_BOLD, fontSize=14)
    h2_style    = ParagraphStyle("h2", parent=styles["Heading2"],
                                 fontName=PDF_FONT_BOLD, fontSize=11)
    body_style  = ParagraphStyle("body", parent=styles["Normal"],
                                 fontName=PDF_FONT, fontSize=8, leading=12)

    story = []
    lookback = 30

    def _header(text):
        story.append(Paragraph(text, h2_style))
        story.append(Spacer(1, 0.3*cm))

    def _table(data, col_widths=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#ADD8E6")),
            ("FONTNAME",    (0, 0), (-1, 0), PDF_FONT_BOLD),
            ("FONTNAME",    (0, 1), (-1, -1), PDF_FONT),
            ("FONTSIZE",    (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.grey),
            ("VALIGN",      (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP",    (0, 0), (-1, -1), True),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

    # 報告標題
    story.append(Paragraph(
        f"廠商情報週報 — {TODAY}（搜尋區間：過去 {lookback} 天）", title_style
    ))
    story.append(Spacer(1, 0.5*cm))

    # 章節一：每週總覽
    _header("一、每週總覽")
    overview_data = [["廠商名稱", "地區", "新品", "停產", "價格異動", "重要公告", "整體動態"]]
    for r in results:
        new_cnt      = len(r.get("new_products", []))
        stopped_cnt  = sum(1 for c in r.get("other_changes", []) if c.get("type") == "停產")
        price_cnt    = sum(1 for c in r.get("other_changes", []) if c.get("type") == "價格調整")
        announce_cnt = sum(1 for c in r.get("other_changes", []) if c.get("type") == "公司公告")
        overview_data.append([
            r.get("vendor", ""), r.get("_region", ""),
            str(new_cnt), str(stopped_cnt), str(price_cnt), str(announce_cnt),
            r.get("activity_level", "無異動"),
        ])
    _table(overview_data, col_widths=[4*cm, 2.5*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm, 2.5*cm])

    story.append(PageBreak())

    # 章節二：新品詳細
    _header("二、新品詳細")
    new_data = [["廠商", "產品名稱", "發布日期", "核實", "類型", "部位", "台灣狀況", "NTD", "通路"]]
    has_products = False
    for r in results:
        for p in r.get("new_products", []):
            new_data.append([
                r.get("vendor", ""),
                (p.get("name") or "")[:30],
                p.get("release_date") or "",
                (p.get("verification") or "❓")[:4],
                (p.get("product_type") or "")[:8],
                (p.get("body_part") or "")[:8],
                p.get("taiwan_status") or "",
                str(p.get("taiwan_price_ntd") or ""),
                (p.get("channel") or "")[:10],
            ])
            has_products = True
    if not has_products:
        new_data.append(["本週無新品發布"] + [""] * 8)
    _table(new_data, col_widths=[2.5*cm, 5*cm, 2.5*cm, 2*cm, 2.5*cm, 2.5*cm, 2*cm, 1.5*cm, 3*cm])

    story.append(PageBreak())

    # 章節三：其他異動
    _header("三、其他異動")
    other_data = [["廠商", "異動類型", "產品/項目", "異動說明", "日期"]]
    has_changes = False
    for r in results:
        for c in r.get("other_changes", []):
            other_data.append([
                r.get("vendor", ""),
                c.get("type", ""),
                c.get("item", "")[:20],
                c.get("description", "")[:60],
                c.get("change_date", ""),
            ])
            has_changes = True
    if not has_changes:
        other_data.append(["本週無其他異動"] + [""] * 4)
    _table(other_data, col_widths=[3*cm, 3*cm, 4*cm, 10*cm, 3*cm])

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(PDF_FONT, 8)
        canvas.drawString(1.5*cm, 1*cm,
                          f"廠商情報週報 {TODAY} | 搜尋區間：過去 {lookback} 天")
        canvas.drawCentredString(landscape(A4)[0] / 2, 1*cm,
                                 f"第 {doc.page} 頁")
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)


# ── 主流程 ──

def main():
    print(f"[廠商報告] 開始執行，日期：{TODAY}")

    vendor_names, vendor_map, _ = load_vendor_config()
    print(f"[廠商報告] 共 {len(vendor_names)} 家廠商：{', '.join(vendor_names)}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"[廠商報告] 輸出目錄：{OUTPUT_DIR}", flush=True)

    # ── 步驟 1：抓取快照 ──
    print("\n[廠商報告] 步驟1：抓取官網快照 ...")
    import snapshot_fetcher
    snap_results = snapshot_fetcher.run_all()

    # ── 步驟 2：逐家分析 ──
    print("\n[廠商報告] 步驟2：分析各廠商異動 ...")
    results = []
    errors  = []
    web_search_count = 0

    for i, name in enumerate(vendor_names):
        info = vendor_map.get(name, {"name": name})
        vid  = info.get("id", name.lower())
        print(f"  → [{name}] ({i+1}/{len(vendor_names)})", flush=True)

        snap_ok = snap_results.get(vid, False)

        if snap_ok:
            # 快照路徑：比較前後兩份
            previous, current = get_latest_snapshots(vid)
            if not previous:
                # 首次執行，只有一份快照，無法比較
                print(f"    首次快照，無前次資料可比較 → 記錄「無異動（首次建立基準）」")
                result = {
                    "vendor": name,
                    "new_products": [],
                    "other_changes": [],
                    "activity_level": "無異動",
                    "_note": "首次建立快照基準，下次執行起開始比較差異",
                }
            else:
                diff_text = compute_diff(previous, current)
                if not diff_text.strip():
                    print(f"    頁面無變化 → 無異動（不呼叫 API）")
                    result = {
                        "vendor": name,
                        "new_products": [],
                        "other_changes": [],
                        "activity_level": "無異動",
                    }
                else:
                    diff_lines = len(diff_text.splitlines())
                    print(f"    偵測到 {diff_lines} 行差異，送 Claude 分析 ...", flush=True)
                    result = analyze_diff(name, diff_text)
        else:
            # 快照失敗 → web_search 備援
            print(f"    快照失敗，改用 web_search 備援 ...", flush=True)
            result = search_vendor(name, info)
            web_search_count += 1
            if web_search_count < len(vendor_names) and i < len(vendor_names) - 1:
                print(f"    [等待 15 秒，避免 rate limit]", flush=True)
                time.sleep(15)

        result["_region"] = info.get("region", "")
        results.append(result)
        if "_error" in result:
            errors.append(f"{name}：{result['_error']}")

        # ── 費用上限守衛：達上限後，用現有資料繼續輸出報告，剩餘廠商標記未分析 ──
        if _claude_cost_state["limit_reached"]:
            remaining = vendor_names[i + 1:]
            if remaining:
                print(
                    f"\n  [費用上限] Claude 費用已達上限，停止分析剩餘 {len(remaining)} 家廠商："
                    f" {', '.join(remaining)}",
                    flush=True,
                )
                print(
                    f"  → 以現有資料繼續輸出報告，未分析廠商將標記為「⚠️ 未分析（Claude費用上限）」",
                    flush=True,
                )
                for skipped in remaining:
                    skipped_info = vendor_map.get(skipped, {})
                    results.append({
                        "vendor":         skipped,
                        "new_products":   [],
                        "other_changes":  [],
                        "activity_level": "⚠️ 未分析",
                        "_region":        skipped_info.get("region", ""),
                        "_error":         "Claude費用上限已達，本廠商未執行分析",
                    })
                    errors.append(f"{skipped}：Claude費用上限已達，跳過")
            break  # 跳出廠商迴圈，繼續往下執行報告生成

    xlsx_path = os.path.join(OUTPUT_DIR, f"vendor_report_{TODAY}.xlsx")
    pdf_path  = os.path.join(OUTPUT_DIR, f"vendor_report_{TODAY}.pdf")

    print("\n[廠商報告] 生成 Excel ...")
    build_excel(results, xlsx_path)

    # ── 步驟 3：生成 HTML 互動分析報告 ──
    print("\n[廠商報告] 載入歷史報告資料 ...")
    history = load_history_data()

    trend_data, alerts = build_trend_data(results, history)
    matrix_data        = build_matrix_data(results, history)
    pricing_data       = build_pricing_data(results, history)

    html_path = os.path.join(OUTPUT_DIR, f"vendor_report_{TODAY}.html")
    print("[廠商報告] 生成 HTML 分析報告 ...", flush=True)
    build_html(html_path, trend_data, matrix_data, pricing_data, results)

    print("[廠商報告] 生成 PDF ...")
    build_pdf(results, pdf_path)

    # 活躍度預警（供 bot.py 解析）
    if alerts:
        print(f"[ALERT] 活躍度上升廠商：{', '.join(alerts)}", flush=True)

    # 執行摘要
    total_new     = sum(len(r.get("new_products", [])) for r in results)
    suspicious    = sum(
        1 for r in results
        for p in r.get("new_products", [])
        if "疑似非新品" in p.get("verification", "")
    )
    total_changes = sum(len(r.get("other_changes", [])) for r in results)
    snap_success  = sum(1 for v in snap_results.values() if v)

    print("\n════════ 執行摘要 ════════")
    print(f"快照抓取成功：{snap_success}/{len(vendor_names)} 家")
    print(f"web_search 備援：{web_search_count} 家")
    print(f"本週發現新品總數：{total_new}")
    print(f"其中標記「疑似非新品」：{suspicious}")
    print(f"其他異動總數：{total_changes}")
    print(f"報告路徑：{OUTPUT_DIR}")
    print(f"HTML 分析報告：{html_path}")
    if errors:
        print(f"\n發生錯誤（{len(errors)} 家）：")
        for e in errors:
            print(f"  • {e}")

    # ── Claude 備援費用摘要 ──
    cs = _claude_cost_state
    print("")
    if cs["call_count"] > 0:
        print(f"── Claude 備援費用摘要 ──")
        print(f"呼叫次數：{cs['call_count']}　輸入 tokens：{cs['total_input_tokens']:,}　輸出 tokens：{cs['total_output_tokens']:,}")
        print(f"估計費用：${cs['total_cost_usd']:.4f} USD　上限：${cs['limit_usd']:.2f} USD")
        if cs["limit_reached"]:
            print(f"⚠️  費用上限已達，部分廠商未執行分析（已標記於報告中）")
    else:
        print(f"Claude 備援：未觸發（Gemini 配額正常）")
    print("══════════════════════════")


if __name__ == "__main__":
    main()
